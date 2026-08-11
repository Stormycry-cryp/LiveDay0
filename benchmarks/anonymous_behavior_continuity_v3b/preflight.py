#!/usr/bin/env python3
"""Fail-closed schema, quota, and anonymity preflight for v3b source archives.

Source identifiers exist only as in-memory dictionary keys. The JSON result contains
counts, schema names, archive hashes, and anonymity statistics; never identifiers or
raw observations.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import statistics
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable


SOURCE_SPECS = {
    "oulad": {
        "archive": "oulad.zip",
        "quota": 120,
        "doi": "10.24432/C5KK69",
    },
    "online_retail_ii": {
        "archive": "retail.zip",
        "quota": 300,
        "doi": "10.24432/C5CG6D",
    },
    "electricity_load_diagrams": {
        "archive": "electricity.zip",
        "quota": 180,
        "doi": "10.24432/C58C86",
    },
}

SLOTS = 20
MIN_SPAN_DAYS = 90
MIN_K = 5


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def slot_index(value: int, start: int, end: int) -> int:
    if end <= start:
        return 0
    return min(SLOTS - 1, ((value - start) * SLOTS) // (end - start + 1))


def sequence_privacy(signatures: Iterable[tuple[str, ...]], quota: int) -> dict[str, object]:
    counts = Counter(signatures)
    k_groups = {signature: count for signature, count in counts.items() if count >= MIN_K}
    eligible_entities = sum(k_groups.values())
    half_sample_capacity = sum(count // 2 for count in k_groups.values())

    # With positional tokens, set-Jaccard >= .90 means at most one of 20 slots differs.
    masked = defaultdict(set)
    for signature in k_groups:
        for index in range(SLOTS):
            key = signature[:index] + ("*",) + signature[index + 1 :]
            masked[key].add(signature)
    near_duplicate_components = sum(1 for values in masked.values() if len(values) > 1)
    max_near_component = max((len(values) for values in masked.values()), default=1)

    sizes = sorted(k_groups.values())
    return {
        "candidate_signature_count": len(counts),
        "k_anonymous_signature_count": len(k_groups),
        "k_anonymous_entity_count": eligible_entities,
        "minimum_k": min(sizes, default=0),
        "median_k": statistics.median(sizes) if sizes else 0,
        "maximum_k": max(sizes, default=0),
        "quota": quota,
        "privacy_capacity_at_max_50pct_per_signature": half_sample_capacity,
        "quota_pass": half_sample_capacity >= quota,
        "membership_inference_prior_ceiling": 0.5 if half_sample_capacity >= quota else None,
        "near_duplicate_mask_groups": near_duplicate_components,
        "maximum_near_duplicate_signature_component": max_near_component,
    }


def find_member(archive: zipfile.ZipFile, basename: str) -> str:
    matches = [name for name in archive.namelist() if Path(name).name.lower() == basename.lower()]
    if len(matches) != 1:
        raise ValueError(f"expected one {basename!r} member, found {matches}")
    return matches[0]


def activity_family(activity_type: str) -> str:
    value = activity_type.strip().lower()
    if value in {"quiz", "questionnaire"}:
        return "assessment"
    if value in {"forumng", "oucollaborate"}:
        return "communication"
    if value in {
        "book",
        "folder",
        "glossary",
        "homepage",
        "oucontent",
        "page",
        "resource",
        "sharedsubpage",
        "subpage",
        "url",
    }:
        return "content"
    return "other"


def oulad_preflight(path: Path) -> dict[str, object]:
    with zipfile.ZipFile(path) as archive:
        student_info_name = find_member(archive, "studentInfo.csv")
        student_vle_name = find_member(archive, "studentVle.csv")
        vle_name = find_member(archive, "vle.csv")

        adult_keys: set[tuple[str, str, str]] = set()
        with archive.open(student_info_name) as raw:
            rows = csv.DictReader(line.decode("utf-8-sig") for line in raw)
            info_schema = rows.fieldnames or []
            for row in rows:
                if row["age_band"] in {"35-55", "55<="}:
                    adult_keys.add((row["code_module"], row["code_presentation"], row["id_student"]))

        site_types: dict[tuple[str, str, str], str] = {}
        with archive.open(vle_name) as raw:
            rows = csv.DictReader(line.decode("utf-8-sig") for line in raw)
            vle_schema = rows.fieldnames or []
            for row in rows:
                site_types[(row["code_module"], row["code_presentation"], row["id_site"])] = activity_family(
                    row["activity_type"]
                )

        daily: dict[str, dict[int, Counter[str]]] = defaultdict(lambda: defaultdict(Counter))
        source_rows = 0
        retained_rows = 0
        with archive.open(student_vle_name) as raw:
            rows = csv.DictReader(line.decode("utf-8-sig") for line in raw)
            student_vle_schema = rows.fieldnames or []
            for row in rows:
                source_rows += 1
                participant_key = (row["code_module"], row["code_presentation"], row["id_student"])
                if participant_key not in adult_keys:
                    continue
                presentation = row["code_presentation"]
                year = int(presentation[:4])
                month = 2 if presentation.endswith("B") else 10
                ordinal = (date(year, month, 1) + timedelta(days=int(row["date"]))).toordinal()
                family = site_types.get((row["code_module"], presentation, row["id_site"]), "other")
                daily[row["id_student"]][ordinal][family] += int(row["sum_click"])
                retained_rows += 1

    signatures: list[tuple[str, ...]] = []
    base_eligible = 0
    for participant_daily in daily.values():
        days = sorted(participant_daily)
        if not days or days[-1] - days[0] < MIN_SPAN_DAYS:
            continue
        slots = [Counter() for _ in range(SLOTS)]
        for day, families in participant_daily.items():
            slots[slot_index(day, days[0], days[-1])].update(families)
        if any(not values for values in slots):
            continue
        base_eligible += 1
        tokens = []
        for values in slots:
            family, clicks = max(values.items(), key=lambda item: (item[1], item[0]))
            bucket = "low" if clicks <= 4 else "mid" if clicks <= 19 else "high"
            tokens.append(f"{family}:{bucket}")
        signatures.append(tuple(tokens))

    return {
        "schema": {
            "studentInfo": info_schema,
            "studentVle": student_vle_schema,
            "vle": vle_schema,
        },
        "source_rows": source_rows,
        "retained_35_plus_rows": retained_rows,
        "ephemeral_35_plus_source_keys": len(adult_keys),
        "base_eligible_entities_90d_20windows": base_eligible,
        "privacy": sequence_privacy(signatures, SOURCE_SPECS["oulad"]["quota"]),
    }


def retail_preflight(path: Path) -> dict[str, object]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - environment error
        raise RuntimeError("openpyxl is required for Online Retail II preflight") from exc

    with zipfile.ZipFile(path) as archive:
        member = find_member(archive, "online_retail_II.xlsx")
        with archive.open(member) as xlsx_handle:
            workbook = openpyxl.load_workbook(xlsx_handle, read_only=True, data_only=True)
            sheets = workbook.sheetnames
            daily: dict[str, dict[int, list[int]]] = defaultdict(lambda: defaultdict(lambda: [0, 0]))
            source_rows = 0
            usable_rows = 0
            schemas: dict[str, list[str]] = {}
            for sheet_name in sheets:
                sheet = workbook[sheet_name]
                rows = sheet.iter_rows(values_only=True)
                schema = [str(value) if value is not None else "" for value in next(rows)]
                schemas[sheet_name] = schema
                index = {name: position for position, name in enumerate(schema)}
                required = {"Invoice", "InvoiceDate", "Customer ID"}
                if not required.issubset(index):
                    raise ValueError(f"sheet {sheet_name!r} missing {sorted(required - set(index))}")
                for row in rows:
                    source_rows += 1
                    customer = row[index["Customer ID"]]
                    timestamp = row[index["InvoiceDate"]]
                    invoice = row[index["Invoice"]]
                    if customer is None or not isinstance(timestamp, (datetime, date)) or invoice is None:
                        continue
                    customer_key = str(customer)
                    ordinal = timestamp.toordinal()
                    cancelled = str(invoice).lower().startswith("c")
                    daily[customer_key][ordinal][1 if cancelled else 0] += 1
                    usable_rows += 1
            workbook.close()

    signatures: list[tuple[str, ...]] = []
    base_eligible = 0
    for participant_daily in daily.values():
        days = sorted(participant_daily)
        if not days or days[-1] - days[0] < MIN_SPAN_DAYS:
            continue
        slots = [[0, 0] for _ in range(SLOTS)]
        for day, values in participant_daily.items():
            slot = slots[slot_index(day, days[0], days[-1])]
            slot[0] += values[0]
            slot[1] += values[1]
        if any(sum(values) == 0 for values in slots):
            continue
        base_eligible += 1
        tokens = []
        for purchases, cancellations in slots:
            state = "mixed" if purchases and cancellations else "cancel" if cancellations else "purchase"
            activity = "low" if purchases + cancellations <= 5 else "high"
            tokens.append(f"{state}:{activity}")
        signatures.append(tuple(tokens))

    return {
        "schema": schemas,
        "source_rows": source_rows,
        "usable_customer_rows": usable_rows,
        "ephemeral_customer_count": len(daily),
        "base_eligible_entities_90d_20windows": base_eligible,
        "privacy": sequence_privacy(signatures, SOURCE_SPECS["online_retail_ii"]["quota"]),
    }


def parse_decimal(value: bytes) -> float:
    return float(value.replace(b",", b"."))


def electricity_preflight(path: Path) -> dict[str, object]:
    with zipfile.ZipFile(path) as archive:
        member = find_member(archive, "LD2011_2014.txt")
        first_nonzero: list[int | None] = []
        last_nonzero: list[int | None] = []
        schema: list[str] = []
        row_count = 0
        with archive.open(member) as raw:
            header = raw.readline().decode("utf-8-sig").strip().split(";")
            schema = header
            client_count = len(header) - 1
            first_nonzero = [None] * client_count
            last_nonzero = [None] * client_count
            for row_index, line in enumerate(raw):
                fields = line.rstrip(b"\r\n").split(b";")
                if len(fields) != len(header):
                    raise ValueError(f"electricity row {row_index + 2} has {len(fields)} fields")
                for index, field in enumerate(fields[1:]):
                    if field and parse_decimal(field) != 0.0:
                        if first_nonzero[index] is None:
                            first_nonzero[index] = row_index
                        last_nonzero[index] = row_index
                row_count += 1

        sums = [[0.0] * SLOTS for _ in first_nonzero]
        counts = [[0] * SLOTS for _ in first_nonzero]
        with archive.open(member) as raw:
            raw.readline()
            for row_index, line in enumerate(raw):
                fields = line.rstrip(b"\r\n").split(b";")
                for index, field in enumerate(fields[1:]):
                    start = first_nonzero[index]
                    end = last_nonzero[index]
                    if start is None or end is None or not (start <= row_index <= end):
                        continue
                    slot = slot_index(row_index, start, end)
                    sums[index][slot] += parse_decimal(field) if field else 0.0
                    counts[index][slot] += 1

    signatures: list[tuple[str, ...]] = []
    base_eligible = 0
    min_span_rows = MIN_SPAN_DAYS * 96
    for index, start in enumerate(first_nonzero):
        end = last_nonzero[index]
        if start is None or end is None or end - start < min_span_rows or any(value == 0 for value in counts[index]):
            continue
        means = [sums[index][slot] / counts[index][slot] for slot in range(SLOTS)]
        baseline = statistics.median(means)
        if baseline <= 0:
            continue
        base_eligible += 1
        tokens = []
        for mean in means:
            state = "low" if mean < baseline * 0.75 else "high" if mean > baseline * 1.25 else "mid"
            tokens.append(state)
        signatures.append(tuple(tokens))

    return {
        "schema": {
            "timestamp_column": schema[0] if schema else None,
            "anonymous_client_columns": max(0, len(schema) - 1),
            "retained_columns": ["relative_slot", "load_state_relative_to_entity_median"],
        },
        "source_rows": row_count,
        "ephemeral_client_count": len(first_nonzero),
        "base_eligible_entities_90d_20windows": base_eligible,
        "privacy": sequence_privacy(signatures, SOURCE_SPECS["electricity_load_diagrams"]["quota"]),
    }


def load_metadata(metadata_dir: Path, name: str) -> dict[str, object]:
    path = metadata_dir / f"{name}-metadata.json"
    payload = json.loads(path.read_text(encoding="utf-8"))
    data = payload["data"]
    return {
        "uci_id": data.get("uci_id"),
        "name": data.get("name"),
        "dataset_doi": data.get("dataset_doi"),
        "repository_url": data.get("repository_url"),
        "last_updated": data.get("last_updated"),
        "metadata_sha256": sha256_file(path),
    }


def run(source_dir: Path, metadata_dir: Path) -> dict[str, object]:
    processors = {
        "oulad": oulad_preflight,
        "online_retail_ii": retail_preflight,
        "electricity_load_diagrams": electricity_preflight,
    }
    results: dict[str, object] = {}
    all_pass = True
    for name, spec in SOURCE_SPECS.items():
        archive = source_dir / spec["archive"]
        metadata_name = "retail" if name == "online_retail_ii" else "electricity" if name == "electricity_load_diagrams" else name
        metadata = load_metadata(metadata_dir, metadata_name)
        if metadata["dataset_doi"] != spec["doi"]:
            raise ValueError(f"{name}: DOI mismatch {metadata['dataset_doi']!r}")
        result = processors[name](archive)
        result["archive"] = {
            "byte_size": archive.stat().st_size,
            "sha256": sha256_file(archive),
            "zip_members": sorted(zipfile.ZipFile(archive).namelist()),
        }
        result["metadata"] = metadata
        result["quota"] = spec["quota"]
        result["preflight_pass"] = bool(result["privacy"]["quota_pass"])
        all_pass = all_pass and bool(result["preflight_pass"])
        results[name] = result
    return {
        "schema_version": "anonymous-behavior-preflight-v3b",
        "rules_frozen_before_source_inspection": {
            "slots": SLOTS,
            "minimum_span_days": MIN_SPAN_DAYS,
            "minimum_exact_signature_k": MIN_K,
            "maximum_selected_fraction_per_signature": 0.5,
            "near_duplicate_threshold": 0.9,
        },
        "sources": results,
        "all_sources_pass": all_pass,
        "freeze_allowed": all_pass,
        "raw_or_source_identifiers_in_output": False,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--metadata-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    result = run(args.source_dir, args.metadata_dir)
    rendered = json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True) + "\n"
    if args.output:
        args.output.write_text(rendered, encoding="utf-8")
    else:
        print(rendered, end="")


if __name__ == "__main__":
    main()
