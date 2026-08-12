#!/usr/bin/env python3
"""Build the privacy-capped v3b Pilot from the three authorized archives.

Raw source identifiers are held only in process memory. Frozen identifiers and
query anchors are independently random and no source-to-release mapping is
written.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import secrets
import statistics
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable
from uuid import uuid4

from benchmarks.anonymous_behavior_continuity_v3b.pilot import (
    SLOTS,
    assign_cluster_splits,
    audit_release,
    near_duplicate_components,
    select_private_members,
)
from benchmarks.anonymous_behavior_continuity_v3b.preflight import (
    activity_family,
    find_member,
    parse_decimal,
    sha256_file,
    slot_index,
)


ROOT = Path(__file__).resolve().parent
ARCHIVE_SPECS = {
    "oulad": {
        "archive": "oulad.zip",
        "sha256": "f2ed1902616c1fe8d2824d872c0b7d2d72be435bf0124d077044fe4be2c6d3e4",
        "cap": 100,
    },
    "online_retail_ii": {
        "archive": "retail.zip",
        "sha256": "572e36277c2390fbfde10664750731e0a86f55e33470d91919085f0408e67bfb",
        "cap": 2,
    },
    "electricity_load_diagrams": {
        "archive": "electricity.zip",
        "sha256": "f6c4d0e0df12ecdb9ea008dd6eef3518adb52c559d04a9bac2e1b81dcfc8d4e1",
        "cap": 115,
    },
}
OUTPUT_FILES = (
    "entities.jsonl",
    "observations.jsonl",
    "cases_test.jsonl",
    "cases_heldout.sealed.jsonl",
    "cases_synthetic_test.jsonl",
    "release_audit.json",
    "manifest.json",
)


@dataclass(frozen=True)
class Candidate:
    source_key: Any
    signature: tuple[str, ...]


def jsonl_bytes(rows: Iterable[dict[str, Any]]) -> bytes:
    return ("\n".join(json.dumps(row, ensure_ascii=False, sort_keys=True) for row in rows) + "\n").encode("utf-8")


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _candidate_groups(candidates: Iterable[Candidate]) -> dict[tuple[str, ...], list[Candidate]]:
    groups: dict[tuple[str, ...], list[Candidate]] = defaultdict(list)
    for candidate in candidates:
        groups[candidate.signature].append(candidate)
    return groups


def extract_oulad(path: Path) -> list[Candidate]:
    with zipfile.ZipFile(path) as archive:
        student_info_name = find_member(archive, "studentInfo.csv")
        student_vle_name = find_member(archive, "studentVle.csv")
        vle_name = find_member(archive, "vle.csv")
        adult_keys: set[tuple[str, str, str]] = set()
        with archive.open(student_info_name) as raw:
            rows = csv.DictReader(line.decode("utf-8-sig") for line in raw)
            for row in rows:
                if row["age_band"] in {"35-55", "55<="}:
                    adult_keys.add((row["code_module"], row["code_presentation"], row["id_student"]))
        site_types: dict[tuple[str, str, str], str] = {}
        with archive.open(vle_name) as raw:
            rows = csv.DictReader(line.decode("utf-8-sig") for line in raw)
            for row in rows:
                site_types[(row["code_module"], row["code_presentation"], row["id_site"])] = activity_family(
                    row["activity_type"]
                )
        daily: dict[str, dict[int, Counter[str]]] = defaultdict(lambda: defaultdict(Counter))
        with archive.open(student_vle_name) as raw:
            rows = csv.DictReader(line.decode("utf-8-sig") for line in raw)
            for row in rows:
                key = (row["code_module"], row["code_presentation"], row["id_student"])
                if key not in adult_keys:
                    continue
                presentation = row["code_presentation"]
                year = int(presentation[:4])
                month = 2 if presentation.endswith("B") else 10
                ordinal = (date(year, month, 1) + timedelta(days=int(row["date"]))).toordinal()
                family = site_types.get((row["code_module"], presentation, row["id_site"]), "other")
                daily[row["id_student"]][ordinal][family] += int(row["sum_click"])

    candidates = []
    for source_key, participant_daily in daily.items():
        days = sorted(participant_daily)
        if not days or days[-1] - days[0] < 90:
            continue
        slots = [Counter() for _ in range(SLOTS)]
        for day, families in participant_daily.items():
            slots[slot_index(day, days[0], days[-1])].update(families)
        if any(not values for values in slots):
            continue
        tokens = []
        for values in slots:
            family, clicks = max(values.items(), key=lambda item: (item[1], item[0]))
            bucket = "low" if clicks <= 4 else "mid" if clicks <= 19 else "high"
            tokens.append(f"{family}:{bucket}")
        candidates.append(Candidate(source_key, tuple(tokens)))
    return candidates


def extract_retail(path: Path) -> list[Candidate]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Online Retail II") from exc

    with zipfile.ZipFile(path) as archive:
        member = find_member(archive, "online_retail_II.xlsx")
        with archive.open(member) as xlsx_handle:
            workbook = openpyxl.load_workbook(xlsx_handle, read_only=True, data_only=True)
            daily: dict[str, dict[int, list[int]]] = defaultdict(lambda: defaultdict(lambda: [0, 0]))
            for sheet_name in workbook.sheetnames:
                rows = workbook[sheet_name].iter_rows(values_only=True)
                schema = [str(value) if value is not None else "" for value in next(rows)]
                index = {name: position for position, name in enumerate(schema)}
                for row in rows:
                    customer = row[index["Customer ID"]]
                    timestamp = row[index["InvoiceDate"]]
                    invoice = row[index["Invoice"]]
                    if customer is None or not isinstance(timestamp, (datetime, date)) or invoice is None:
                        continue
                    daily[str(customer)][timestamp.toordinal()][str(invoice).lower().startswith("c")] += 1
            workbook.close()

    candidates = []
    for source_key, participant_daily in daily.items():
        days = sorted(participant_daily)
        if not days or days[-1] - days[0] < 90:
            continue
        slots = [[0, 0] for _ in range(SLOTS)]
        for day, values in participant_daily.items():
            target = slots[slot_index(day, days[0], days[-1])]
            target[0] += values[0]
            target[1] += values[1]
        if any(sum(values) == 0 for values in slots):
            continue
        tokens = []
        for purchases, cancellations in slots:
            state = "mixed" if purchases and cancellations else "cancel" if cancellations else "purchase"
            activity = "low" if purchases + cancellations <= 5 else "high"
            tokens.append(f"{state}:{activity}")
        candidates.append(Candidate(source_key, tuple(tokens)))
    return candidates


def extract_electricity(path: Path) -> list[Candidate]:
    with zipfile.ZipFile(path) as archive:
        member = find_member(archive, "LD2011_2014.txt")
        with archive.open(member) as raw:
            header = raw.readline().decode("utf-8-sig").strip().split(";")
            first: list[int | None] = [None] * (len(header) - 1)
            last: list[int | None] = [None] * (len(header) - 1)
            for row_index, line in enumerate(raw):
                fields = line.rstrip(b"\r\n").split(b";")
                for index, field in enumerate(fields[1:]):
                    if field and parse_decimal(field) != 0.0:
                        if first[index] is None:
                            first[index] = row_index
                        last[index] = row_index
        sums = [[0.0] * SLOTS for _ in first]
        counts = [[0] * SLOTS for _ in first]
        with archive.open(member) as raw:
            raw.readline()
            for row_index, line in enumerate(raw):
                fields = line.rstrip(b"\r\n").split(b";")
                for index, field in enumerate(fields[1:]):
                    start = first[index]
                    end = last[index]
                    if start is None or end is None or not (start <= row_index <= end):
                        continue
                    slot = slot_index(row_index, start, end)
                    sums[index][slot] += parse_decimal(field) if field else 0.0
                    counts[index][slot] += 1

    candidates = []
    for source_key, start in enumerate(first):
        end = last[source_key]
        if start is None or end is None or end - start < 90 * 96 or any(value == 0 for value in counts[source_key]):
            continue
        means = [sums[source_key][slot] / counts[source_key][slot] for slot in range(SLOTS)]
        baseline = statistics.median(means)
        if baseline <= 0:
            continue
        tokens = tuple(
            "low" if mean < baseline * 0.75 else "high" if mean > baseline * 1.25 else "mid"
            for mean in means
        )
        candidates.append(Candidate(source_key, tokens))
    return candidates


def _query(language: str, category: str, anchor: str, state: str, qualifier: str = "") -> str:
    templates = {
        "zh": f"匿名实体 {anchor} 的{qualifier}行为状态是什么？检索 {state}。",
        "en": f"What is anonymous entity {anchor}'s {qualifier}behavior state? Retrieve {state}.",
        "es": f"¿Cuál es el estado de comportamiento {qualifier}de la entidad anónima {anchor}? Recupera {state}.",
    }
    return templates[language]


def build_real_cases(
    internal_entities: list[dict[str, Any]], observations: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    observations_by_entity: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for observation in observations:
        observations_by_entity[observation["entity_id"]].append(observation)
    languages = ("zh", "en", "es")
    by_split: dict[str, list[dict[str, Any]]] = {"test": [], "heldout": []}
    language_index = 0
    for entity in sorted(internal_entities, key=lambda row: row["entity_id"]):
        split = entity["split"]
        if split not in by_split:
            continue
        rows = sorted(observations_by_entity[entity["entity_id"]], key=lambda row: row["relative_slot"])
        dominant = Counter(row["behavior_state"] for row in rows).most_common(1)[0][0]
        repeated = [row for row in rows if row["behavior_state"] == dominant][-5:]
        definitions = [
            ("current_shift" if rows[0]["behavior_state"] != rows[-1]["behavior_state"] else "stable_pattern", 19, [rows[-1]], "current "),
            ("repeated_behavior", 19, repeated, "repeated "),
            ("time_segment_continuity", 14, rows[10:15], "third-quarter "),
        ]
        for category, cutoff, expected, qualifier in definitions:
            language = languages[language_index % len(languages)]
            language_index += 1
            future = [row["observation_id"] for row in rows if row["relative_slot"] > cutoff]
            by_split[split].append(
                {
                    "schema_version": "anonymous-behavior-case-v3b-pilot",
                    "case_id": str(uuid4()),
                    "case_origin": "real_derived",
                    "query_origin": "synthetic",
                    "synthetic_role": "query_interface",
                    "evidence_origin": "real_source_derived",
                    "split": split,
                    "entity_id": entity["entity_id"],
                    "source_family": entity["source_family"],
                    "language": language,
                    "category": category,
                    "query": _query(language, category, entity["query_anchor"], expected[-1]["behavior_state"], qualifier),
                    "cutoff_slot": cutoff,
                    "expected_observation_ids": [row["observation_id"] for row in expected],
                    "forbidden_future_observation_ids": future,
                }
            )
    return by_split["test"], by_split["heldout"]


def build_synthetic_cases(test_cases: list[dict[str, Any]]) -> list[dict[str, Any]]:
    categories = (
        "near_duplicate_noise",
        "synthetic_reversal",
        "stale_marker",
        "deletion_propagation",
        "lifecycle_supersession",
    )
    seeds_by_entity: dict[str, dict[str, Any]] = {}
    for case in test_cases:
        seeds_by_entity.setdefault(case["entity_id"], case)
    seeds = list(seeds_by_entity.values())[:10]
    rows = []
    for category in categories:
        for seed in seeds:
            expected = [] if category == "deletion_propagation" else seed["expected_observation_ids"][-1:]
            rows.append(
                {
                    **seed,
                    "case_id": str(uuid4()),
                    "case_origin": "synthetic_mechanics",
                    "evidence_origin": "synthetic",
                    "synthetic_role": category,
                    "category": category,
                    "expected_observation_ids": expected,
                    "target_observation_ids": seed["expected_observation_ids"][-1:],
                }
            )
    return rows


def build(source_dir: Path, output_dir: Path) -> dict[str, Any]:
    existing = [name for name in OUTPUT_FILES if (output_dir / name).exists()]
    if existing:
        raise FileExistsError(f"refusing to overwrite frozen Pilot files: {existing}")
    extractors = {
        "oulad": extract_oulad,
        "online_retail_ii": extract_retail,
        "electricity_load_diagrams": extract_electricity,
    }
    snapshots = {}
    all_candidates: dict[str, list[Candidate]] = {}
    for family, spec in ARCHIVE_SPECS.items():
        archive = source_dir / spec["archive"]
        digest = sha256_file(archive)
        if digest != spec["sha256"]:
            raise ValueError(f"{family}: snapshot hash changed: {digest}")
        snapshots[family] = {"sha256": digest, "bytes": archive.stat().st_size}
        all_candidates[family] = extractors[family](archive)

    rng = secrets.SystemRandom()
    selected_by_family: dict[str, list[tuple[tuple[str, ...], Candidate]]] = {}
    selection_audits = {}
    eligible_counts: dict[str, dict[tuple[str, ...], int]] = {}
    weighted_components = []
    component_members: dict[tuple[str, ...], tuple[str, set[tuple[str, ...]]]] = {}
    for family, candidates in all_candidates.items():
        groups = _candidate_groups(candidates)
        eligible_counts[family] = {signature: len(members) for signature, members in groups.items() if len(members) >= 5}
        selected, selection_audit = select_private_members(groups, cap=ARCHIVE_SPECS[family]["cap"], rng=rng)
        selected_by_family[family] = selected
        selection_audits[family] = selection_audit
        selected_counts = Counter(signature for signature, _ in selected)
        for component in near_duplicate_components(eligible_counts[family]):
            wrapped = {("source=" + family,) + signature for signature in component}
            weight = sum(selected_counts[signature] for signature in component)
            weighted_components.append((wrapped, weight))
            for wrapped_signature in wrapped:
                component_members[wrapped_signature] = (family, component)
    wrapped_assignments = assign_cluster_splits(
        weighted_components,
        target_fractions={"train": 0.5, "test": 0.25, "heldout": 0.25},
    )

    signature_ids: dict[tuple[str, tuple[str, ...]], str] = {}
    cluster_ids: dict[tuple[str, tuple[tuple[str, ...], ...]], str] = {}
    internal_entities = []
    observations = []
    for family, selected in selected_by_family.items():
        for signature, _candidate in selected:
            wrapped = ("source=" + family,) + signature
            component = component_members[wrapped][1]
            component_key = (family, tuple(sorted(component)))
            entity_id = str(uuid4())
            signature_id = signature_ids.setdefault((family, signature), str(uuid4()))
            cluster_id = cluster_ids.setdefault(component_key, str(uuid4()))
            split = wrapped_assignments[wrapped]
            anchor = "anchor-" + uuid4().hex[:12]
            internal_entities.append(
                {
                    "entity_id": entity_id,
                    "source_family": family,
                    "split": split,
                    "query_anchor": anchor,
                    "signature_group_id": signature_id,
                    "near_duplicate_cluster_id": cluster_id,
                    "signature": list(signature),
                }
            )
            for slot, state in enumerate(signature):
                observations.append(
                    {
                        "schema_version": "anonymous-behavior-observation-v3b-pilot",
                        "observation_id": str(uuid4()),
                        "entity_id": entity_id,
                        "source_family": family,
                        "split": split,
                        "relative_slot": slot,
                        "temporal_quartile": slot // 5,
                        "behavior_state": state,
                        "evidence_origin": "real_source_derived",
                    }
                )

    test_cases, heldout_cases = build_real_cases(internal_entities, observations)
    synthetic_cases = build_synthetic_cases(test_cases)
    release_audit = audit_release(internal_entities, observations, test_cases + heldout_cases, eligible_counts)
    release_audit.update(
        {
            "schema_version": "anonymous-behavior-release-audit-v3b-pilot",
            "selection": selection_audits,
            "actual": {
                "real_entities": len(internal_entities),
                "real_observations": len(observations),
                "real_derived_test_cases": len(test_cases),
                "real_derived_heldout_cases": len(heldout_cases),
                "synthetic_test_cases": len(synthetic_cases),
            },
        }
    )
    if not release_audit["passed"]:
        raise ValueError(f"release audit failed: {release_audit['errors']}")

    public_entities = [
        {
            "schema_version": "anonymous-behavior-entity-v3b-pilot",
            "entity_id": row["entity_id"],
            "source_family": row["source_family"],
            "split": row["split"],
            "query_anchor": row["query_anchor"],
            "query_anchor_origin": "synthetic_query_interface",
            "signature_group_id": row["signature_group_id"],
            "near_duplicate_cluster_id": row["near_duplicate_cluster_id"],
            "observation_count": SLOTS,
            "minimum_source_span_days": 90,
            "source_identifier_retained": False,
        }
        for row in internal_entities
    ]
    artifacts = {
        "entities.jsonl": jsonl_bytes(public_entities),
        "observations.jsonl": jsonl_bytes(observations),
        "cases_test.jsonl": jsonl_bytes(test_cases),
        "cases_heldout.sealed.jsonl": jsonl_bytes(heldout_cases),
        "cases_synthetic_test.jsonl": jsonl_bytes(synthetic_cases),
        "release_audit.json": (json.dumps(release_audit, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode("utf-8"),
    }
    manifest = {
        "schema_version": "anonymous-behavior-manifest-v3b-pilot",
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source_snapshots": snapshots,
        "real_layer": {
            "entities": len(public_entities),
            "observations": len(observations),
            "cases_test": len(test_cases),
            "cases_heldout_sealed": len(heldout_cases),
        },
        "synthetic_layer": {
            "entities": 0,
            "observations": 0,
            "test_cases": len(synthetic_cases),
            "query_interface_cases": len(test_cases) + len(heldout_cases),
        },
        "distribution": {
            "entities_by_source": dict(sorted(Counter(row["source_family"] for row in public_entities).items())),
            "entities_by_split": dict(sorted(Counter(row["split"] for row in public_entities).items())),
            "real_cases_by_split": {"test": len(test_cases), "heldout": len(heldout_cases)},
            "real_cases_by_language": dict(sorted(Counter(row["language"] for row in test_cases + heldout_cases).items())),
            "real_cases_by_category": dict(sorted(Counter(row["category"] for row in test_cases + heldout_cases).items())),
            "synthetic_cases_by_category": dict(sorted(Counter(row["category"] for row in synthetic_cases).items())),
        },
        "heldout_boundary": {
            "status": "sealed_not_run",
            "case_file": "cases_heldout.sealed.jsonl",
        },
        "artifacts": {name: {"sha256": sha256_bytes(value), "bytes": len(value)} for name, value in artifacts.items()},
    }
    artifacts["manifest.json"] = (json.dumps(manifest, ensure_ascii=False, indent=2, sort_keys=True) + "\n").encode("utf-8")
    output_dir.mkdir(parents=True, exist_ok=True)
    for name, value in artifacts.items():
        (output_dir / name).write_bytes(value)
    return manifest


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, default=ROOT)
    args = parser.parse_args()
    print(json.dumps(build(args.source_dir, args.output_dir), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
