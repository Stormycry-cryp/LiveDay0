#!/usr/bin/env python3
"""Aggregate-only, fail-closed preflight helpers for v3c.

This module never accepts v3b heldout cases. Cross-version exclusion is derived
only from the public v3b entity and observation releases.
"""

from __future__ import annotations

import json
import hashlib
import argparse
import csv
import statistics
import sys
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence


SLOTS = 20
NEAR_DUPLICATE_MAX_DISTANCE = 1
SPLITS = ("train", "test", "heldout")
MIN_ENTITIES_PER_SPLIT = 30
MIN_COMPONENTS_PER_EVALUATION_SPLIT = 6
MAX_COMPONENT_FRACTION = 0.35
MIN_SOURCE_FAMILIES = 2
MAX_SOURCE_FRACTION = 0.8
CASE_CUTOFF_SLOT = 14
MIN_K = 5
MIN_REAL_ENTITIES = 90
MIN_REAL_OBSERVATIONS = 1800

ROOT = Path(__file__).resolve().parent
V3B_ROOT = ROOT.parent / "anonymous_behavior_continuity_v3b"
SOURCE_SPECS = {
    "oulad": {
        "archive": "oulad.zip",
        "bytes": 46748244,
        "sha256": "f2ed1902616c1fe8d2824d872c0b7d2d72be435bf0124d077044fe4be2c6d3e4",
        "doi": "10.24432/C5KK69",
        "uci_id": 349,
    },
    "online_retail_ii": {
        "archive": "retail.zip",
        "bytes": 45622418,
        "sha256": "572e36277c2390fbfde10664750731e0a86f55e33470d91919085f0408e67bfb",
        "doi": "10.24432/C5CG6D",
        "uci_id": 502,
    },
    "electricity_load_diagrams": {
        "archive": "electricity.zip",
        "bytes": 261335609,
        "sha256": "f6c4d0e0df12ecdb9ea008dd6eef3518adb52c559d04a9bac2e1b81dcfc8d4e1",
        "doi": "10.24432/C58C86",
        "uci_id": 321,
    },
}
OFFICIAL_SOURCE_EVIDENCE = {
    "oulad": {
        "doi_pass": True,
        "license_pass": True,
        "metadata_sha256": "079799e66dec01a31b211709750f48667da88476551258393bbfef2def49c874",
        "official_page_sha256": "cbb00ea10a0adbe206f1dfc156155c73b25abcbde10567dbc51d906c914bd898",
    },
    "online_retail_ii": {
        "doi_pass": True,
        "license_pass": True,
        "metadata_sha256": "ada567e65f3d9481518ad7228da51860c2d5f45dfa6c7efebee8196853735c78",
        "official_page_sha256": "34d60d8ec9f688fba780bf9bfbf27c685b3698f2d144bd66ac053b34d5721536",
    },
    "electricity_load_diagrams": {
        "doi_pass": True,
        "license_pass": True,
        "metadata_sha256": "b2dcc4da3b2bb104f2bab064bdf5f1da280fb2446a1a9d464f4319963023f56f",
        "official_page_sha256": "7ad7902222452534703d76ffdf69ab9bdba571dab35360ec6bdf1085edad79dd",
    },
}

REQUIRED_AGGREGATE_FIELDS = {
    "all_snapshot_identities_pass",
    "all_schema_and_license_checks_pass",
    "v3b_public_signatures_reconstructed",
    "cross_version_excluded_entities_by_source",
    "base_eligible_entities_by_source",
    "v3c_projected_entities_by_source",
    "exact_k_eligible_entities_by_source",
    "privacy_capacity_by_source",
    "case_eligible_entities_by_source",
    "near_duplicate_components_by_source",
    "component_split_capacity",
    "mia_reidentification_feasibility",
    "freeze_allowed",
    "errors",
}
FORBIDDEN_KEY_FRAGMENTS = {
    "source_entity",
    "source_id",
    "student_id",
    "customer_id",
    "client_id",
    "case_id",
    "query",
    "expected_id",
    "forbidden_future_id",
    "raw_observation",
    "raw_curve",
    "exact_timestamp",
}


@dataclass(frozen=True)
class Candidate:
    source_key: Any
    v3b_signature: tuple[str, ...]
    v3c_signature: tuple[str, ...]


def slot_index(value: int, start: int, end: int) -> int:
    if end <= start:
        return 0
    return min(SLOTS - 1, ((value - start) * SLOTS) // (end - start + 1))


def find_member(archive: zipfile.ZipFile, basename: str) -> str:
    matches = [name for name in archive.namelist() if Path(name).name.lower() == basename.lower()]
    if len(matches) != 1:
        raise ValueError(f"expected one {basename!r} member, found {matches}")
    return matches[0]


def activity_family(activity_type: str) -> str:
    value = activity_type.strip().lower()
    if value in {"quiz", "questionnaire"}:
        return "assessment"
    if value in {"forumng", "oucollaborate"}:
        return "communication"
    if value in {
        "book", "folder", "glossary", "homepage", "oucontent", "page",
        "resource", "sharedsubpage", "subpage", "url",
    }:
        return "content"
    return "other"


def parse_decimal(value: bytes) -> float:
    return float(value.replace(b",", b"."))


def signature_distance(left: Sequence[str], right: Sequence[str]) -> int:
    if len(left) != SLOTS or len(right) != SLOTS:
        raise ValueError(f"signatures must contain exactly {SLOTS} slots")
    return sum(a != b for a, b in zip(left, right, strict=True))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def verify_official_evidence(evidence_dir: Path) -> dict[str, dict[str, Any]]:
    """Verify official metadata identity and license link, aggregate only."""

    output: dict[str, dict[str, Any]] = {}
    for source, spec in SOURCE_SPECS.items():
        uci_id = spec["uci_id"]
        metadata_path = evidence_dir / f"metadata-{uci_id}.json"
        page_path = evidence_dir / f"page-{uci_id}.html"
        data = json.loads(metadata_path.read_text(encoding="utf-8")).get("data", {})
        page = page_path.read_text(encoding="utf-8")
        output[source] = {
            "doi_pass": data.get("uci_id") == uci_id and data.get("dataset_doi") == spec["doi"],
            "license_pass": "https://creativecommons.org/licenses/by/4.0/legalcode" in page,
            "metadata_sha256": sha256_file(metadata_path),
            "official_page_sha256": sha256_file(page_path),
        }
    return output


def verified_parse(
    path: Path,
    *,
    expected_bytes: int,
    expected_sha256: str,
    parser: Any,
) -> Any:
    """Verify the complete immutable snapshot before invoking its parser."""

    if not path.is_file() or path.stat().st_size != expected_bytes or sha256_file(path) != expected_sha256:
        raise ValueError(f"snapshot identity mismatch: {path.name}")
    return parser(path)


def is_case_eligible(signature: Sequence[str], *, cutoff_slot: int = CASE_CUTOFF_SLOT) -> bool:
    """Require real recurrence and a bounded transition before the fixed cutoff."""

    if len(signature) != SLOTS or cutoff_slot not in range(SLOTS - 1):
        raise ValueError("invalid case eligibility signature or cutoff")
    visible = tuple(signature[: cutoff_slot + 1])
    quartiles_by_state: dict[str, set[int]] = defaultdict(set)
    for slot, state in enumerate(visible):
        quartiles_by_state[state].add(slot // 5)
    recurrence = any(len(quartiles) >= 3 for quartiles in quartiles_by_state.values())
    transition = any(left != right for left, right in zip(visible, visible[1:], strict=False))
    return recurrence and transition


def audit_private_case_capacity(groups: Mapping[tuple[str, ...], Sequence[Any]]) -> dict[str, int]:
    case_eligible = {signature: list(members) for signature, members in groups.items() if is_case_eligible(signature)}
    exact_k = {signature: members for signature, members in case_eligible.items() if len(members) >= MIN_K}
    return {
        "projected_entities": sum(len(members) for members in groups.values()),
        "case_eligible_entities": sum(len(members) for members in case_eligible.values()),
        "exact_k_eligible_entities": sum(len(members) for members in exact_k.values()),
        "exact_k_eligible_signatures": len(exact_k),
        "privacy_capacity": sum(len(members) // 2 for members in exact_k.values()),
    }


def audit_reidentification_feasibility(groups: Mapping[tuple[str, ...], Sequence[Any]]) -> dict[str, Any]:
    eligible = {
        signature: len(members)
        for signature, members in groups.items()
        if is_case_eligible(signature) and len(members) >= MIN_K
    }
    minimum_exact = min(eligible.values(), default=0)
    minimum_leave_one_out: int | None = None
    for signature in eligible:
        for slot in range(SLOTS):
            matches = sum(
                count
                for candidate, count in eligible.items()
                if candidate[:slot] + candidate[slot + 1 :] == signature[:slot] + signature[slot + 1 :]
            )
            minimum_leave_one_out = matches if minimum_leave_one_out is None else min(minimum_leave_one_out, matches)
    leave_one_out = minimum_leave_one_out or 0
    passed = minimum_exact >= MIN_K and leave_one_out >= MIN_K
    return {
        "passed": passed,
        "minimum_exact_signature_matches": minimum_exact,
        "minimum_leave_one_slot_out_matches": leave_one_out,
        "unique_reidentification_feasible": not passed,
    }


def reconstruct_public_signatures(
    entities_path: Path,
    observations_path: Path,
) -> dict[str, set[tuple[str, ...]]]:
    """Reconstruct source-local v3b signatures without reading any case file."""

    entities: dict[str, str] = {}
    for line in entities_path.read_text(encoding="utf-8").splitlines():
        if not line:
            continue
        row = json.loads(line)
        entity_id = row["entity_id"]
        if entity_id in entities:
            raise ValueError("duplicate public v3b entity_id")
        entities[entity_id] = row["source_family"]

    slots: dict[str, dict[int, str]] = defaultdict(dict)
    for line in observations_path.read_text(encoding="utf-8").splitlines():
        if not line:
            continue
        row = json.loads(line)
        entity_id = row["entity_id"]
        if entity_id not in entities or row["source_family"] != entities[entity_id]:
            raise ValueError("public v3b observation entity/source mismatch")
        slot = row["relative_slot"]
        if slot not in range(SLOTS) or slot in slots[entity_id]:
            raise ValueError("invalid or duplicate public v3b relative slot")
        slots[entity_id][slot] = row["behavior_state"]

    output: dict[str, set[tuple[str, ...]]] = defaultdict(set)
    for entity_id, family in entities.items():
        if set(slots[entity_id]) != set(range(SLOTS)):
            raise ValueError("public v3b entity does not have all 20 slots")
        output[family].add(tuple(slots[entity_id][slot] for slot in range(SLOTS)))
    return dict(output)


def is_cross_version_excluded(
    source_family: str,
    signature: Sequence[str],
    released_by_source: Mapping[str, set[tuple[str, ...]]],
) -> bool:
    return any(
        signature_distance(signature, released) <= NEAR_DUPLICATE_MAX_DISTANCE
        for released in released_by_source.get(source_family, set())
    )


def signature_components(signatures: Iterable[tuple[str, ...]]) -> list[set[tuple[str, ...]]]:
    """Return transitive components connected by exact/one-slot distance."""

    values = sorted(set(signatures))
    parent = list(range(len(values)))

    def root(index: int) -> int:
        while parent[index] != index:
            parent[index] = parent[parent[index]]
            index = parent[index]
        return index

    def union(left: int, right: int) -> None:
        a, b = root(left), root(right)
        if a != b:
            parent[b] = a

    masked: dict[tuple[str, ...], int] = {}
    for index, signature in enumerate(values):
        if len(signature) != SLOTS:
            raise ValueError(f"signature must contain {SLOTS} slots")
        for slot in range(SLOTS):
            key = signature[:slot] + ("*",) + signature[slot + 1 :]
            if key in masked:
                union(index, masked[key])
            else:
                masked[key] = index

    components: dict[int, set[tuple[str, ...]]] = defaultdict(set)
    for index, signature in enumerate(values):
        components[root(index)].add(signature)
    return sorted(components.values(), key=lambda value: min(value))


def exclude_released_components(
    source_family: str,
    groups: Mapping[tuple[str, ...], Sequence[Any]],
    released_by_source: Mapping[str, set[tuple[str, ...]]],
) -> tuple[dict[tuple[str, ...], list[Any]], dict[str, int]]:
    components = signature_components(groups)
    excluded = [
        component
        for component in components
        if any(is_cross_version_excluded(source_family, signature, released_by_source) for signature in component)
    ]
    excluded_signatures = set().union(*excluded) if excluded else set()
    kept = {signature: list(members) for signature, members in groups.items() if signature not in excluded_signatures}
    return kept, {
        "candidate_entities": sum(len(members) for members in groups.values()),
        "candidate_signatures": len(groups),
        "excluded_components": len(excluded),
        "excluded_signatures": len(excluded_signatures),
        "excluded_entities": sum(len(groups[signature]) for signature in excluded_signatures),
        "remaining_entities": sum(len(members) for members in kept.values()),
    }


def audit_split_capacity(assignments: Mapping[str, Sequence[Mapping[str, str]]]) -> dict[str, Any]:
    errors: list[str] = []
    details: dict[str, Any] = {}
    for split in SPLITS:
        rows = list(assignments.get(split, []))
        source_counts = Counter(row["source_family"] for row in rows)
        component_counts = Counter(row["component"] for row in rows)
        total = len(rows)
        if total < MIN_ENTITIES_PER_SPLIT:
            errors.append(f"{split}: fewer than {MIN_ENTITIES_PER_SPLIT} entities")
        if split in {"test", "heldout"}:
            if len(component_counts) < MIN_COMPONENTS_PER_EVALUATION_SPLIT:
                errors.append(f"{split}: fewer than {MIN_COMPONENTS_PER_EVALUATION_SPLIT} components")
            if total and max(component_counts.values(), default=0) / total > MAX_COMPONENT_FRACTION:
                errors.append(f"{split}: component fraction above {MAX_COMPONENT_FRACTION}")
            if len(source_counts) < MIN_SOURCE_FAMILIES:
                errors.append(f"{split}: fewer than {MIN_SOURCE_FAMILIES} sources")
            if total and max(source_counts.values(), default=0) / total > MAX_SOURCE_FRACTION:
                errors.append(f"{split}: source fraction above {MAX_SOURCE_FRACTION}")
        details[split] = {
            "entities": total,
            "components": len(component_counts),
            "sources": len(source_counts),
            "maximum_component_fraction": round(max(component_counts.values(), default=0) / total, 6)
            if total
            else 0.0,
            "maximum_source_fraction": round(max(source_counts.values(), default=0) / total, 6) if total else 0.0,
        }
    return {"passed": not errors, "errors": errors, "splits": details}


def validate_aggregate_output(result: Mapping[str, Any]) -> None:
    missing = REQUIRED_AGGREGATE_FIELDS - set(result)
    if missing:
        raise ValueError(f"missing aggregate output fields: {sorted(missing)}")

    def walk(value: Any) -> None:
        if isinstance(value, Mapping):
            for key, nested in value.items():
                normalized = str(key).lower().replace("-", "_")
                if any(fragment in normalized for fragment in FORBIDDEN_KEY_FRAGMENTS):
                    raise ValueError(f"forbidden aggregate output key: {key}")
                walk(nested)
        elif isinstance(value, list):
            for nested in value:
                walk(nested)

    walk(result)


def extract_oulad(path: Path) -> tuple[list[Candidate], dict[str, Any]]:
    with zipfile.ZipFile(path) as archive:
        info_name = find_member(archive, "studentInfo.csv")
        student_vle_name = find_member(archive, "studentVle.csv")
        vle_name = find_member(archive, "vle.csv")
        with archive.open(info_name) as raw:
            rows = csv.DictReader(line.decode("utf-8-sig") for line in raw)
            info_schema = rows.fieldnames or []
            adult_keys = {
                (row["code_module"], row["code_presentation"], row["id_student"])
                for row in rows
                if row["age_band"] in {"35-55", "55<="}
            }
        with archive.open(vle_name) as raw:
            rows = csv.DictReader(line.decode("utf-8-sig") for line in raw)
            vle_schema = rows.fieldnames or []
            site_types = {
                (row["code_module"], row["code_presentation"], row["id_site"]): activity_family(row["activity_type"])
                for row in rows
            }
        daily: dict[str, dict[int, Counter[str]]] = defaultdict(lambda: defaultdict(Counter))
        source_rows = retained_rows = 0
        with archive.open(student_vle_name) as raw:
            rows = csv.DictReader(line.decode("utf-8-sig") for line in raw)
            student_vle_schema = rows.fieldnames or []
            for row in rows:
                source_rows += 1
                key = (row["code_module"], row["code_presentation"], row["id_student"])
                if key not in adult_keys:
                    continue
                presentation = row["code_presentation"]
                ordinal = (
                    date(int(presentation[:4]), 2 if presentation.endswith("B") else 10, 1)
                    + timedelta(days=int(row["date"]))
                ).toordinal()
                family = site_types.get((row["code_module"], presentation, row["id_site"]), "other")
                daily[row["id_student"]][ordinal][family] += int(row["sum_click"])
                retained_rows += 1

    candidates: list[Candidate] = []
    for source_key, values_by_day in daily.items():
        days = sorted(values_by_day)
        if not days or days[-1] - days[0] < 90:
            continue
        slots = [Counter() for _ in range(SLOTS)]
        for day, values in values_by_day.items():
            slots[slot_index(day, days[0], days[-1])].update(values)
        if any(not values for values in slots):
            continue
        old_tokens = []
        new_tokens = []
        for values in slots:
            family, clicks = max(values.items(), key=lambda item: (item[1], item[0]))
            old_tokens.append(f"{family}:{'low' if clicks <= 4 else 'mid' if clicks <= 19 else 'high'}")
            new_tokens.append(family)
        candidates.append(Candidate(source_key, tuple(old_tokens), tuple(new_tokens)))
    return candidates, {
        "schema_pass": {"age_band", "id_student"}.issubset(info_schema)
        and {"id_site", "activity_type"}.issubset(vle_schema)
        and {"id_student", "id_site", "date", "sum_click"}.issubset(student_vle_schema),
        "source_rows": source_rows,
        "privacy_filter_rows": retained_rows,
        "base_eligible_entities": len(candidates),
    }


def extract_retail(path: Path) -> tuple[list[Candidate], dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl 3.1+ is required for the official Retail snapshot") from exc

    with zipfile.ZipFile(path) as archive:
        member = find_member(archive, "online_retail_II.xlsx")
        with archive.open(member) as handle:
            workbook = openpyxl.load_workbook(handle, read_only=True, data_only=True)
            daily: dict[str, dict[int, list[int]]] = defaultdict(lambda: defaultdict(lambda: [0, 0]))
            source_rows = usable_rows = 0
            schema_pass = True
            for sheet_name in workbook.sheetnames:
                rows = workbook[sheet_name].iter_rows(values_only=True)
                schema = [str(value) if value is not None else "" for value in next(rows)]
                index = {name: i for i, name in enumerate(schema)}
                required = {"Invoice", "InvoiceDate", "Customer ID"}
                schema_pass = schema_pass and required.issubset(index)
                if not required.issubset(index):
                    raise ValueError(f"Retail sheet {sheet_name!r} missing schema")
                for row in rows:
                    source_rows += 1
                    customer = row[index["Customer ID"]]
                    timestamp = row[index["InvoiceDate"]]
                    invoice = row[index["Invoice"]]
                    if customer is None or not isinstance(timestamp, (datetime, date)) or invoice is None:
                        continue
                    daily[str(customer)][timestamp.toordinal()][str(invoice).lower().startswith("c")] += 1
                    usable_rows += 1
            workbook.close()

    candidates = []
    for source_key, values_by_day in daily.items():
        days = sorted(values_by_day)
        if not days or days[-1] - days[0] < 90:
            continue
        slots = [[0, 0] for _ in range(SLOTS)]
        for day, values in values_by_day.items():
            target = slots[slot_index(day, days[0], days[-1])]
            target[0] += values[0]
            target[1] += values[1]
        if any(sum(values) == 0 for values in slots):
            continue
        old_tokens = []
        new_tokens = []
        for purchases, cancellations in slots:
            state = "mixed" if purchases and cancellations else "cancel" if cancellations else "purchase"
            old_tokens.append(f"{state}:{'low' if purchases + cancellations <= 5 else 'high'}")
            new_tokens.append(state)
        candidates.append(Candidate(source_key, tuple(old_tokens), tuple(new_tokens)))
    return candidates, {
        "schema_pass": schema_pass,
        "source_rows": source_rows,
        "privacy_filter_rows": usable_rows,
        "base_eligible_entities": len(candidates),
    }


def extract_electricity(path: Path) -> tuple[list[Candidate], dict[str, Any]]:
    with zipfile.ZipFile(path) as archive:
        member = find_member(archive, "LD2011_2014.txt")
        with archive.open(member) as raw:
            header = raw.readline().decode("utf-8-sig").strip().split(";")
            first: list[int | None] = [None] * (len(header) - 1)
            last: list[int | None] = [None] * (len(header) - 1)
            source_rows = 0
            for row_index, line in enumerate(raw):
                source_rows += 1
                fields = line.rstrip(b"\r\n").split(b";")
                for index, field in enumerate(fields[1:]):
                    if field and parse_decimal(field) != 0:
                        if first[index] is None:
                            first[index] = row_index
                        last[index] = row_index
        sums = [[0.0] * SLOTS for _ in first]
        counts = [[0] * SLOTS for _ in first]
        with archive.open(member) as raw:
            raw.readline()
            for row_index, line in enumerate(raw):
                fields = line.rstrip(b"\r\n").split(b";")
                for index, field in enumerate(fields[1:]):
                    start, end = first[index], last[index]
                    if start is None or end is None or not (start <= row_index <= end):
                        continue
                    slot = slot_index(row_index, start, end)
                    sums[index][slot] += parse_decimal(field) if field else 0.0
                    counts[index][slot] += 1

    candidates = []
    for source_key, start in enumerate(first):
        end = last[source_key]
        if start is None or end is None or end - start < 90 * 96 or any(value == 0 for value in counts[source_key]):
            continue
        means = [sums[source_key][slot] / counts[source_key][slot] for slot in range(SLOTS)]
        median = statistics.median(means)
        if median <= 0:
            continue
        old = tuple("low" if mean < median * 0.75 else "high" if mean > median * 1.25 else "mid" for mean in means)
        new = tuple("below_median" if mean < median else "at_or_above_median" for mean in means)
        candidates.append(Candidate(source_key, old, new))
    return candidates, {
        "schema_pass": len(header) == 371,
        "source_rows": source_rows,
        "privacy_filter_rows": source_rows,
        "base_eligible_entities": len(candidates),
    }


def group_candidates(candidates: Sequence[Candidate], attribute: str) -> dict[tuple[str, ...], list[Candidate]]:
    groups: dict[tuple[str, ...], list[Candidate]] = defaultdict(list)
    for candidate in candidates:
        groups[getattr(candidate, attribute)].append(candidate)
    return dict(groups)


def assign_component_capacity(
    components_by_source: Mapping[str, Sequence[tuple[set[tuple[str, ...]], int]]],
) -> dict[str, Any]:
    """Find a bounded whole-component split; identities never leave the process.

    A component's weight is capacity, not a requirement to publish every
    available member.  The contract permits selecting *at most* half of each
    eligible signature, so this feasibility check may take a smaller bounded
    count from an assigned component.
    """

    items = [
        (source, component_index, weight)
        for source, components in components_by_source.items()
        for component_index, (_component, weight) in enumerate(components)
        if weight
    ]
    total = sum(weight for _source, _index, weight in items)
    ordered = sorted(items, key=lambda item: (item[0], item[1]))

    def find_evaluation_plan(
        available: Sequence[tuple[str, int, int]],
        *,
        preserve_another_evaluation: bool = False,
    ) -> list[tuple[str, int, int]] | None:
        sources = tuple(sorted(components_by_source))
        source_index = {source: index for index, source in enumerate(sources)}
        available_components = Counter(source for source, _index, _capacity in available)
        # state: (total, per-source entities, per-source components) -> plan
        states: dict[
            tuple[int, tuple[int, ...], tuple[int, ...]], tuple[tuple[str, int, int], ...]
        ] = {
            (0, (0,) * len(sources), (0,) * len(sources)): ()
        }
        for source, index, capacity in available:
            next_states = dict(states)
            for (selected_total, source_counts, source_components), selected in states.items():
                maximum = min(
                    capacity,
                    int(MIN_ENTITIES_PER_SPLIT * MAX_COMPONENT_FRACTION),
                    MIN_ENTITIES_PER_SPLIT - selected_total,
                    int(MIN_ENTITIES_PER_SPLIT * MAX_SOURCE_FRACTION)
                    - source_counts[source_index[source]],
                )
                for take in range(1, maximum + 1):
                    counts = list(source_counts)
                    counts[source_index[source]] += take
                    components = list(source_components)
                    components[source_index[source]] += 1
                    key = (selected_total + take, tuple(counts), tuple(components))
                    next_states.setdefault(key, selected + ((source, index, take),))
            states = next_states
        candidates = [
            selected
            for (selected_total, source_counts, source_components), selected in states.items()
            if selected_total == MIN_ENTITIES_PER_SPLIT
            and sum(source_components) == MIN_COMPONENTS_PER_EVALUATION_SPLIT
            and sum(value > 0 for value in source_counts) >= MIN_SOURCE_FAMILIES
            and (
                not preserve_another_evaluation
                or (
                    sum(available_components.values()) - sum(source_components)
                    >= MIN_COMPONENTS_PER_EVALUATION_SPLIT
                    and sum(
                        available_components[source] - source_components[source_index[source]] > 0
                        for source in sources
                    )
                    >= MIN_SOURCE_FAMILIES
                )
            )
        ]
        return list(min(candidates, key=lambda value: tuple(value))) if candidates else None

    evaluation_plans: tuple[list[tuple[str, int, int]], list[tuple[str, int, int]]] | None = None
    if ordered:
        # Rotations prevent the first valid evaluation plan from consuming the
        # only components needed by the second split.  This is a feasibility
        # search over aggregate component capacities, never source identities.
        for offset in range(len(ordered)):
            rotated = ordered[offset:] + ordered[:offset]
            test_plan = find_evaluation_plan(rotated, preserve_another_evaluation=True)
            if test_plan is None:
                continue
            used = {(source, index) for source, index, _take in test_plan}
            heldout_plan = find_evaluation_plan([item for item in rotated if item[:2] not in used])
            if heldout_plan is not None:
                used_pair = used | {(source, index) for source, index, _take in heldout_plan}
                train_capacity = sum(capacity for source, index, capacity in rotated if (source, index) not in used_pair)
                if train_capacity < MIN_ENTITIES_PER_SPLIT:
                    continue
                evaluation_plans = (test_plan, heldout_plan)
                break

    assignments: dict[str, list[dict[str, str]]] = {split: [] for split in SPLITS}
    used_components: set[tuple[str, int]] = set()
    if evaluation_plans is not None:
        for split, plan in zip(("test", "heldout"), evaluation_plans, strict=True):
            for source, index, take in plan:
                used_components.add((source, index))
                assignments[split].extend(
                    {"source_family": source, "component": f"component-{source}-{index}"}
                    for _ in range(take)
                )
        remaining_train = MIN_ENTITIES_PER_SPLIT
        for source, index, capacity in ordered:
            if (source, index) in used_components or not remaining_train:
                continue
            take = min(capacity, remaining_train)
            assignments["train"].extend(
                {"source_family": source, "component": f"component-{source}-{index}"} for _ in range(take)
            )
            remaining_train -= take
    audit = audit_split_capacity(assignments)
    audit["total_privacy_capacity"] = total
    audit["feasible_selected_entities"] = sum(len(rows) for rows in assignments.values())
    return audit


def run_preflight(source_dir: Path) -> dict[str, Any]:
    released = reconstruct_public_signatures(V3B_ROOT / "entities.jsonl", V3B_ROOT / "observations.jsonl")
    official_evidence = OFFICIAL_SOURCE_EVIDENCE
    extractors = {
        "oulad": extract_oulad,
        "online_retail_ii": extract_retail,
        "electricity_load_diagrams": extract_electricity,
    }
    snapshot_identities: dict[str, dict[str, Any]] = {}
    source_schema: dict[str, dict[str, Any]] = {}
    cross_version: dict[str, dict[str, int]] = {}
    base_counts: dict[str, int] = {}
    projected_counts: dict[str, int] = {}
    exact_k_counts: dict[str, int] = {}
    privacy_capacity: dict[str, int] = {}
    case_eligible_counts: dict[str, int] = {}
    component_counts: dict[str, int] = {}
    components_by_source: dict[str, list[tuple[set[tuple[str, ...]], int]]] = {}
    reidentification_by_source: dict[str, dict[str, Any]] = {}

    # All byte identities are verified before the first source parser runs.
    for source, spec in SOURCE_SPECS.items():
        path = source_dir / spec["archive"]
        if not path.is_file() or path.stat().st_size != spec["bytes"] or sha256_file(path) != spec["sha256"]:
            raise ValueError(f"snapshot identity mismatch: {spec['archive']}")
        with zipfile.ZipFile(path) as archive:
            if archive.testzip() is not None:
                raise ValueError(f"snapshot ZIP integrity failure: {spec['archive']}")
        snapshot_identities[source] = {
            "sha256": spec["sha256"],
            "bytes": spec["bytes"],
            "zip_integrity_pass": True,
            "doi": spec["doi"],
            "license": "CC BY 4.0",
        }

    for source, spec in SOURCE_SPECS.items():
        candidates, schema = extractors[source](source_dir / spec["archive"])
        source_schema[source] = schema
        base_counts[source] = len(candidates)
        old_groups = group_candidates(candidates, "v3b_signature")
        kept_old, exclusion = exclude_released_components(source, old_groups, released)
        kept_source_keys = {candidate.source_key for members in kept_old.values() for candidate in members}
        remaining = [candidate for candidate in candidates if candidate.source_key in kept_source_keys]
        cross_version[source] = exclusion
        new_groups = group_candidates(remaining, "v3c_signature")
        capacity = audit_private_case_capacity(new_groups)
        reidentification_by_source[source] = audit_reidentification_feasibility(new_groups)
        projected_counts[source] = capacity["projected_entities"]
        exact_k_counts[source] = capacity["exact_k_eligible_entities"]
        privacy_capacity[source] = capacity["privacy_capacity"]
        case_eligible_counts[source] = capacity["case_eligible_entities"]
        eligible_groups = {
            signature: members
            for signature, members in new_groups.items()
            if is_case_eligible(signature) and len(members) >= MIN_K
        }
        components = signature_components(eligible_groups)
        component_counts[source] = len(components)
        components_by_source[source] = [
            (component, sum(len(eligible_groups[signature]) // 2 for signature in component)) for component in components
        ]

    split_capacity = assign_component_capacity(components_by_source)
    eligible_reid = [value for source, value in reidentification_by_source.items() if privacy_capacity[source] > 0]
    minimum_exact = min((value["minimum_exact_signature_matches"] for value in eligible_reid), default=0)
    minimum_leave_one_out = min((value["minimum_leave_one_slot_out_matches"] for value in eligible_reid), default=0)
    mia_reid = {
        "passed": bool(
            sum(privacy_capacity.values()) >= MIN_REAL_ENTITIES
            and minimum_exact >= MIN_K
            and minimum_leave_one_out >= MIN_K
        ),
        "maximum_signature_membership_posterior": 0.5,
        "minimum_exact_signature_matches": minimum_exact,
        "minimum_leave_one_slot_out_matches": minimum_leave_one_out,
        "unique_reidentification_feasible": not bool(
            sum(privacy_capacity.values()) >= MIN_REAL_ENTITIES
            and minimum_exact >= MIN_K
            and minimum_leave_one_out >= MIN_K
        ),
        "by_source": reidentification_by_source,
        "note": "full empirical MIA/re-identification runs after deterministic selection at M2",
    }
    errors = []
    if not all(value["schema_pass"] for value in source_schema.values()):
        errors.append("schema gate failed")
    if not all(value["doi_pass"] and value["license_pass"] for value in official_evidence.values()):
        errors.append("official DOI/license evidence gate failed")
    if sum(privacy_capacity.values()) < MIN_REAL_ENTITIES:
        errors.append("real entity privacy capacity below 90")
    if sum(privacy_capacity.values()) * SLOTS < MIN_REAL_OBSERVATIONS:
        errors.append("real observation privacy capacity below 1800")
    if sum(value > 0 for value in privacy_capacity.values()) < MIN_SOURCE_FAMILIES:
        errors.append("fewer than two source families have privacy capacity")
    if not split_capacity["passed"]:
        errors.extend(split_capacity["errors"])
    if not mia_reid["passed"]:
        errors.append("MIA/re-identification feasibility gate failed")
    result = {
        "schema_version": "anonymous-behavior-preflight-result-v3c",
        "all_snapshot_identities_pass": True,
        "all_schema_and_license_checks_pass": all(value["schema_pass"] for value in source_schema.values())
        and all(value["doi_pass"] and value["license_pass"] for value in official_evidence.values()),
        "snapshot_identities": snapshot_identities,
        "official_source_evidence": official_evidence,
        "source_schema_audit": source_schema,
        "v3b_public_signatures_reconstructed": sum(len(values) for values in released.values()),
        "cross_version_excluded_entities_by_source": {
            source: audit["excluded_entities"] for source, audit in cross_version.items()
        },
        "cross_version_exclusion_audit_by_source": cross_version,
        "base_eligible_entities_by_source": base_counts,
        "v3c_projected_entities_by_source": projected_counts,
        "exact_k_eligible_entities_by_source": exact_k_counts,
        "privacy_capacity_by_source": privacy_capacity,
        "case_eligible_entities_by_source": case_eligible_counts,
        "near_duplicate_components_by_source": component_counts,
        "component_split_capacity": split_capacity,
        "mia_reidentification_feasibility": mia_reid,
        "freeze_allowed": not errors,
        "errors": errors,
    }
    validate_aggregate_output(result)
    return result


def render_report(result: Mapping[str, Any]) -> str:
    lines = [
        "# v3c official-source preflight report",
        "",
        f"Decision: `{'GO_FREEZE' if result['freeze_allowed'] else 'NO_GO_PREFLIGHT'}`.",
        "This report contains aggregate counts only; no source ID, raw observation, case, or heldout row is retained.",
        "",
        "| Source | Base eligible | Cross-version excluded | v3c projected | Case eligible | Exact k eligible | Privacy capacity | Components |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for source in SOURCE_SPECS:
        lines.append(
            f"| {source} | {result['base_eligible_entities_by_source'][source]} | "
            f"{result['cross_version_excluded_entities_by_source'][source]} | "
            f"{result['v3c_projected_entities_by_source'][source]} | "
            f"{result['case_eligible_entities_by_source'][source]} | "
            f"{result['exact_k_eligible_entities_by_source'][source]} | "
            f"{result['privacy_capacity_by_source'][source]} | "
            f"{result['near_duplicate_components_by_source'][source]} |"
        )
    lines.extend(["", "Errors:"])
    lines.extend(f"- {error}" for error in result["errors"])
    if not result["errors"]:
        lines.append("- none")
    lines.extend([
        "",
        "M1 writes no frozen entity, observation, case, label, split manifest, RED result, or heldout artifact.",
        "Synthetic material contributes zero source/entity/observation/case capacity.",
        "",
    ])
    return "\n".join(lines)


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    args = parser.parse_args(argv)
    output_dir = args.output_dir.resolve()
    result_path = output_dir / "preflight_result.json"
    report_path = output_dir / "preflight_report.md"
    if result_path.exists() or report_path.exists():
        parser.error("refusing to overwrite an existing preflight output")
    output_dir.mkdir(parents=True, exist_ok=True)
    try:
        result = run_preflight(args.source_dir.resolve())
    except Exception as exc:
        print(f"preflight failed closed: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 3
    result_path.write_text(json.dumps(result, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    report_path.write_text(render_report(result), encoding="utf-8")
    print(json.dumps({"freeze_allowed": result["freeze_allowed"], "errors": result["errors"]}, sort_keys=True))
    return 0 if result["freeze_allowed"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
